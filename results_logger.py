import os
from datetime import datetime
from typing import Optional

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

RESULTS_FILE = "model_comparison.xlsx"

HEADERS = [
    "Timestamp",
    "Stage",
    "Architecture",
    "Mode",
    "Optimizer",
    "LR",
    "Epochs Trained",
    "Train Acc (%)",
    "Test Acc (%)",
    "Notes",
]

# ---------------------------------------------------------------------------
# Styling helpers
# ---------------------------------------------------------------------------

_HEADER_FILL  = PatternFill("solid", fgColor="1F4E79")
_STAGE_FILLS  = {
    "A": PatternFill("solid", fgColor="D9E1F2"),
    "B": PatternFill("solid", fgColor="E2EFDA"),
    "C": PatternFill("solid", fgColor="FFF2CC"),
}
_THIN_BORDER  = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin"),
)
_COL_WIDTHS   = [20, 8, 18, 20, 12, 10, 16, 15, 15, 35]


def _apply_header_style(ws) -> None:
    for col_idx, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font      = Font(bold=True, color="FFFFFF", size=11)
        cell.fill      = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = _THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = _COL_WIDTHS[col_idx - 1]
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"


def _style_data_row(ws, row: int, stage: str) -> None:
    fill = _STAGE_FILLS.get(stage, PatternFill())
    for col_idx in range(1, len(HEADERS) + 1):
        cell = ws.cell(row=row, column=col_idx)
        cell.fill      = fill
        cell.border    = _THIN_BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def log_result(
    stage: str,
    architecture: str,
    mode: str,
    optimizer: str,
    lr: float,
    epochs_trained: int,
    train_acc: Optional[float],
    test_acc: Optional[float],
    notes: str = "",
) -> None:
    """
    Append one experiment row to model_comparison.xlsx.
    Creates the file with a styled header if it does not exist yet.

    Parameters
    ----------
    stage          : "A", "B", or "C"
    architecture   : e.g. "VGG_SmallSigmoid", "ResNet-50", "ConvAutoencoder"
    mode           : e.g. "scratch", "frozen backbone", "full fine-tune",
                         "frozen encoder", "pretrain only"
    optimizer      : e.g. "SGD", "Adam"
    lr             : learning rate used
    epochs_trained : actual number of epochs run (after early stopping)
    train_acc      : final training accuracy [0-1], or None if not applicable
    test_acc       : final / best test accuracy [0-1], or None if not applicable
    notes          : any free-text remarks
    """
    if os.path.exists(RESULTS_FILE):
        wb = load_workbook(RESULTS_FILE)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Results"
        _apply_header_style(ws)

    next_row = ws.max_row + 1

    train_acc_pct = round(train_acc * 100, 2) if train_acc is not None else "N/A"
    test_acc_pct  = round(test_acc  * 100, 2) if test_acc  is not None else "N/A"

    row_values = [
        datetime.now().strftime("%Y-%m-%d %H:%M"),
        stage,
        architecture,
        mode,
        optimizer,
        lr,
        epochs_trained,
        train_acc_pct,
        test_acc_pct,
        notes,
    ]

    for col_idx, value in enumerate(row_values, start=1):
        ws.cell(row=next_row, column=col_idx, value=value)

    _style_data_row(ws, next_row, stage)

    wb.save(RESULTS_FILE)
    print(f"  [Excel] Result logged → {RESULTS_FILE}  (row {next_row})")
